from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from .experience_requirements import find_experience_requirement, requirement_label
from .job_filters import is_actionable_job
from .models import JobPosting, ScoreResult, applicant_sort_value


APPLICATION_STATUS_OPTIONS = ["Not Applied Yet", "Applied"]
SOURCE_STATUS_OPTIONS = ["No", "Yes"]
USER_COLUMNS = ["Applied", "Use As Source", "Application Date", "Application Notes"]

RANKED_HEADERS = [
    "Rank",
    "Job ID",
    "Overall Score",
    "ATS Score",
    "Title",
    "Company",
    "Location",
    "Applicants",
    "Applicant Count Text",
    "Application Status",
    "Job Link",
    "Google Resume Link",
    "Google Resume ID",
    "OneDrive Resume Link",
    "OneDrive Resume ID",
    "OneDrive Cover Letter Link",
    "OneDrive Cover Letter ID",
    "OneDrive Cold Outreach Link",
    "OneDrive Cold Outreach ID",
    "Applied",
    "Use As Source",
    "Application Date",
    "Application Notes",
    "Generated Resume",
    "Generated Cover Letter",
    "Cold Outreach",
    "Role Fit",
    "Skill Match",
    "Experience Match",
    "Domain Fit",
    "Seniority/Location",
    "ATS Keyword Coverage",
    "Matched Keywords",
    "Missing Keywords",
    "Best Source Resume",
    "Scraped At",
    "Notes",
]


RAW_HEADERS = [
    "Job ID",
    "Title",
    "Company",
    "Location",
    "Applicants",
    "Applicant Count Text",
    "Application Status",
    "URL",
    "Source URL",
    "Scraped At",
    "Description Preview",
]

EXCLUDED_HEADERS = [
    "Reason",
    "Job ID",
    "Title",
    "Company",
    "Location",
    "Required Experience",
    "Experience Phrase",
    "Application Status",
    "Applicants",
    "Applicant Count Text",
    "URL",
    "Scraped At",
    "Description Preview",
]


def write_excel_report(
    jobs: Iterable[JobPosting],
    scores: dict[str, ScoreResult],
    output_path: str | Path,
    min_score: float,
    application_status: dict[str, dict[str, str]] | None = None,
    config: dict[str, Any] | None = None,
) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. Run: pip install -r job_scanner/requirements.txt") from exc

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    cfg = config or {}
    visible_jobs = [job for job in jobs if _include_in_excel(job, cfg)]
    ranked = sorted(
        (
            job
            for job in visible_jobs
            if job.key() in scores
            and scores[job.key()].overall_score >= min_score
            and is_actionable_job(job, cfg)
        ),
        key=lambda item: _ranked_sort_key(item, scores, cfg),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Ranked Jobs"
    ws.append(RANKED_HEADERS)
    ranked_run_rows: list[int] = []
    current_run = ""
    rank = 0
    for job in ranked:
        run_label = _run_label(job, cfg)
        if run_label != current_run:
            current_run = run_label
            _append_run_header(ws, run_label, len(RANKED_HEADERS))
            ranked_run_rows.append(ws.max_row)
        rank += 1
        score = scores[job.key()]
        preserved = (application_status or {}).get(job.key(), {})
        row = [
            rank,
            job.key(),
            score.overall_score,
            score.resume_ats_score,
            job.title,
            job.company,
            job.location,
            job.applicant_count,
            job.applicant_count_text,
            job.application_status,
            job.url,
            score.google_doc_url,
            score.google_doc_id,
            score.onedrive_doc_url,
            score.onedrive_doc_id,
            score.onedrive_cover_letter_url,
            score.onedrive_cover_letter_id,
            score.onedrive_cold_outreach_url,
            score.onedrive_cold_outreach_id,
            _application_value(preserved.get("Applied", "")),
            _source_value(preserved.get("Use As Source", "")),
            preserved.get("Application Date", ""),
            preserved.get("Application Notes", preserved.get("Notes", "")),
            score.resume_path,
            score.cover_letter_path,
            score.cold_outreach_path,
            score.role_fit,
            score.skill_match,
            score.experience_match,
            score.domain_fit,
            score.seniority_location_fit,
            score.ats_keyword_coverage,
            ", ".join(score.matched_keywords),
            ", ".join(score.missing_keywords),
            score.matched_resume_path,
            job.scraped_at,
            score.notes,
        ]
        ws.append(row)

    raw_ws = wb.create_sheet("Raw Jobs")
    raw_ws.append(RAW_HEADERS)
    raw_run_rows: list[int] = []
    current_raw_run = ""
    raw_jobs = sorted(visible_jobs, key=_scraped_at_sort_key)
    for job in raw_jobs:
        run_label = _run_label(job, cfg)
        if run_label != current_raw_run:
            current_raw_run = run_label
            _append_run_header(raw_ws, run_label, len(RAW_HEADERS))
            raw_run_rows.append(raw_ws.max_row)
        raw_ws.append(
            [
                job.job_id,
                job.title,
                job.company,
                job.location,
                job.applicant_count,
                job.applicant_count_text,
                job.application_status,
                job.url,
                job.source_url,
                job.scraped_at,
                job.description[:1200],
            ]
        )

    excluded_ws = wb.create_sheet("Excluded Jobs")
    excluded_ws.append(EXCLUDED_HEADERS)
    excluded_jobs = sorted(
        [job for job in jobs if _exclusion_reason(job, cfg)],
        key=_scraped_at_sort_key,
    )
    for job in excluded_jobs:
        requirement = find_experience_requirement(job.full_text())
        excluded_ws.append(
            [
                _exclusion_reason(job, cfg),
                job.job_id,
                job.title,
                job.company,
                job.location,
                requirement_label(requirement),
                requirement.phrase if requirement else "",
                job.application_status,
                job.applicant_count,
                job.applicant_count_text,
                job.url,
                job.scraped_at,
                job.description[:1200],
            ]
        )

    _style_sheet(ws, RANKED_HEADERS, Table, TableStyleInfo, ColorScaleRule, Font, PatternFill, Alignment, get_column_letter)
    _style_sheet(raw_ws, RAW_HEADERS, Table, TableStyleInfo, ColorScaleRule, Font, PatternFill, Alignment, get_column_letter)
    _style_sheet(excluded_ws, EXCLUDED_HEADERS, Table, TableStyleInfo, ColorScaleRule, Font, PatternFill, Alignment, get_column_letter)
    _style_run_header_rows(ws, ranked_run_rows, Font, PatternFill, Alignment)
    _style_run_header_rows(raw_ws, raw_run_rows, Font, PatternFill, Alignment)
    _add_ranked_dropdowns(ws, DataValidation)

    for row in range(2, ws.max_row + 1):
        header_cols = {str(cell.value): cell.column for cell in ws[1] if cell.value}
        job_link_col = header_cols.get("Job Link")
        hyperlink_cols = [
            header_cols.get("Google Resume Link"),
            header_cols.get("OneDrive Resume Link"),
            header_cols.get("OneDrive Cover Letter Link"),
            header_cols.get("OneDrive Cold Outreach Link"),
            header_cols.get("Generated Resume"),
            header_cols.get("Generated Cover Letter"),
            header_cols.get("Cold Outreach"),
        ]
        if job_link_col:
            ws.cell(row=row, column=job_link_col).hyperlink = ws.cell(row=row, column=job_link_col).value
            ws.cell(row=row, column=job_link_col).style = "Hyperlink"
        for col in [item for item in hyperlink_cols if item]:
            if ws.cell(row=row, column=col).value:
                ws.cell(row=row, column=col).hyperlink = ws.cell(row=row, column=col).value
                ws.cell(row=row, column=col).style = "Hyperlink"

    raw_header_cols = {str(cell.value): cell.column for cell in raw_ws[1] if cell.value}
    raw_url_col = raw_header_cols.get("URL")
    for row in range(2, raw_ws.max_row + 1):
        if raw_url_col:
            raw_ws.cell(row=row, column=raw_url_col).hyperlink = raw_ws.cell(row=row, column=raw_url_col).value
            raw_ws.cell(row=row, column=raw_url_col).style = "Hyperlink"

    excluded_header_cols = {str(cell.value): cell.column for cell in excluded_ws[1] if cell.value}
    excluded_url_col = excluded_header_cols.get("URL")
    for row in range(2, excluded_ws.max_row + 1):
        if excluded_url_col and excluded_ws.cell(row=row, column=excluded_url_col).value:
            excluded_ws.cell(row=row, column=excluded_url_col).hyperlink = excluded_ws.cell(row=row, column=excluded_url_col).value
            excluded_ws.cell(row=row, column=excluded_url_col).style = "Hyperlink"

    tmp_output = output.with_name(f".{output.stem}.tmp{output.suffix}")
    try:
        wb.save(tmp_output)
        tmp_output.replace(output)
    except OSError as exc:
        try:
            tmp_output.unlink(missing_ok=True)
        except OSError:
            pass
        print(
            f"Excel workbook is locked or unavailable; keeping existing workbook unchanged. "
            f"The next run will retry: {exc}"
        )
    return str(output)


def read_excel_application_status(path: str | Path) -> dict[str, dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. Run: pip install -r job_scanner/requirements.txt") from exc

    workbook_path = Path(path)
    if not workbook_path.exists() or workbook_path.stat().st_size == 0:
        return {}

    try:
        wb = load_workbook(workbook_path, read_only=False, data_only=True, keep_links=False)
    except Exception:
        return {}
    if "Ranked Jobs" not in wb.sheetnames:
        return {}
    ws = wb["Ranked Jobs"]
    headers = {str(cell.value).strip(): index for index, cell in enumerate(ws[1], start=1) if cell.value}
    job_id_col = headers.get("Job ID")
    job_link_col = headers.get("Job Link")
    if not job_id_col and not job_link_col:
        return {}

    result: dict[str, dict[str, str]] = {}
    for row in range(2, ws.max_row + 1):
        job_id = str(ws.cell(row=row, column=job_id_col).value or "").strip() if job_id_col else ""
        if not job_id and job_link_col:
            link = str(ws.cell(row=row, column=job_link_col).value or "").strip()
            job_id = link.rstrip("/").split("/")[-1] if link else ""
        if not job_id:
            continue
        preserved: dict[str, str] = {}
        for header in USER_COLUMNS:
            col = headers.get(header)
            preserved[header] = str(ws.cell(row=row, column=col).value or "").strip() if col else ""
        for header in ["OneDrive Resume ID", "OneDrive Resume Link", "Google Resume ID", "Google Resume Link"]:
            col = headers.get(header)
            if col:
                preserved[header] = str(ws.cell(row=row, column=col).value or "").strip()
        result[job_id] = preserved
    return result


def _include_in_excel(job: JobPosting, config: dict[str, Any]) -> bool:
    if bool(config.get("excel_hide_closed_jobs", True)) and not job.accepting_applications:
        return False
    if _exceeds_required_experience(job, config):
        return False
    return True


def _exclusion_reason(job: JobPosting, config: dict[str, Any]) -> str:
    if bool(config.get("excel_hide_closed_jobs", True)) and not job.accepting_applications:
        return "No longer accepting applications"
    if _exceeds_required_experience(job, config):
        requirement = find_experience_requirement(job.full_text())
        return f"Requires {requirement_label(requirement)} experience"
    return ""


def _exceeds_required_experience(job: JobPosting, config: dict[str, Any]) -> bool:
    requirement = find_experience_requirement(job.full_text())
    if not requirement:
        return False
    max_allowed = float(config.get("max_required_experience_years", 5.99))
    return requirement.minimum_years > max_allowed


def _ranked_sort_key(job: JobPosting, scores: dict[str, ScoreResult], config: dict[str, Any]) -> tuple[Any, ...]:
    sort_mode = str(config.get("excel_sort_mode", "latest_first")).strip().lower()
    if sort_mode in {"applicants", "lowest_applicants"}:
        return (applicant_sort_value(job), -scores[job.key()].overall_score, _scraped_at_sort_key(job))
    return (_scraped_at_sort_key(job), applicant_sort_value(job), -scores[job.key()].overall_score)


def _scraped_at_sort_key(job: JobPosting) -> float:
    value = str(job.scraped_at or "").strip()
    if not value:
        return 0.0
    try:
        return -datetime.fromisoformat(value.replace("Z", "+00:00")).timestamp()
    except ValueError:
        return 0.0


def _append_run_header(ws: Any, label: str, column_count: int) -> None:
    ws.append([label, *[""] * (column_count - 1)])


def _run_label(job: JobPosting, config: dict[str, Any]) -> str:
    run_time = _run_bucket(job, config)
    display_time = run_time.strftime("%B %-d, %Y %-I:%M %p %Z")
    return f"JOBS based on Latest Run as of -- {display_time}"


def _run_bucket(job: JobPosting, config: dict[str, Any]) -> datetime:
    scraped_at = _parse_scraped_at(job.scraped_at)
    schedule = config.get("launch_schedule", {})
    if schedule.get("mode") != "daily_times":
        return scraped_at.replace(minute=0, second=0, microsecond=0)
    times = sorted(_parse_launch_time_value(value) for value in schedule.get("times", []))
    if not times:
        return scraped_at.replace(minute=0, second=0, microsecond=0)

    candidates = [
        scraped_at.replace(hour=hour, minute=minute, second=0, microsecond=0)
        for hour, minute in times
    ]
    prior = [candidate for candidate in candidates if candidate <= scraped_at]
    if prior:
        return max(prior)

    previous_hour, previous_minute = times[-1]
    return (scraped_at - timedelta(days=1)).replace(
        hour=previous_hour,
        minute=previous_minute,
        second=0,
        microsecond=0,
    )


def _parse_scraped_at(value: str) -> datetime:
    text = str(value or "").strip()
    if text:
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
            return parsed.astimezone()
        except ValueError:
            pass
    return datetime.now().astimezone()


def _parse_launch_time_value(value: Any) -> tuple[int, int]:
    if isinstance(value, dict):
        return int(value.get("hour", 8)), int(value.get("minute", 0))
    hour_text, _, minute_text = str(value).partition(":")
    return int(hour_text), int(minute_text or 0)


def _style_sheet(ws, headers, Table, TableStyleInfo, ColorScaleRule, Font, PatternFill, Alignment, get_column_letter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    widths = {
        "A": 8,
        "B": 18,
        "C": 14,
        "D": 12,
        "E": 34,
        "F": 24,
        "G": 22,
        "H": 38,
        "I": 42,
        "J": 28,
        "K": 42,
        "L": 28,
        "M": 18,
        "N": 16,
        "O": 18,
        "P": 32,
        "Q": 42,
    }
    for idx, _header in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = widths.get(letter, 24)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    if ws.max_row >= 2 and ws.max_column >= 2:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=_table_name(ws.title), ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if ws.title == "Ranked Jobs" and ws.max_row >= 2:
        ws.conditional_formatting.add(
            f"C2:C{ws.max_row}",
            ColorScaleRule(start_type="num", start_value=6, start_color="F8696B", mid_type="num", mid_value=7.5, mid_color="FFEB84", end_type="num", end_value=10, end_color="63BE7B"),
        )
        ws.conditional_formatting.add(
            f"D2:D{ws.max_row}",
            ColorScaleRule(start_type="num", start_value=50, start_color="F8696B", mid_type="num", mid_value=75, mid_color="FFEB84", end_type="num", end_value=100, end_color="63BE7B"),
        )


def _style_run_header_rows(ws: Any, row_numbers: list[int], Font: Any, PatternFill: Any, Alignment: Any) -> None:
    fill = PatternFill("solid", fgColor="305496")
    font = Font(color="FFFFFF", bold=True, size=12)
    for row_number in row_numbers:
        ws.row_dimensions[row_number].height = 24
        for cell in ws[row_number]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(wrap_text=True, vertical="center")


def _add_ranked_dropdowns(ws: Any, DataValidation: Any) -> None:
    if ws.title != "Ranked Jobs" or ws.max_row < 2:
        return
    headers = {str(cell.value): cell.column for cell in ws[1] if cell.value}
    applied_col = headers.get("Applied")
    source_col = headers.get("Use As Source")
    if applied_col:
        applied_dv = DataValidation(type="list", formula1='"Not Applied Yet,Applied"', allow_blank=False)
        applied_dv.error = "Choose Applied or Not Applied Yet."
        applied_dv.errorTitle = "Invalid applied status"
        ws.add_data_validation(applied_dv)
        applied_dv.add(f"{ws.cell(2, applied_col).coordinate}:{ws.cell(ws.max_row, applied_col).coordinate}")
    if source_col:
        source_dv = DataValidation(type="list", formula1='"No,Yes"', allow_blank=False)
        source_dv.error = "Choose Yes or No."
        source_dv.errorTitle = "Invalid source status"
        ws.add_data_validation(source_dv)
        source_dv.add(f"{ws.cell(2, source_col).coordinate}:{ws.cell(ws.max_row, source_col).coordinate}")


def _application_value(value: str) -> str:
    return "Applied" if str(value).strip().lower() == "applied" else "Not Applied Yet"


def _source_value(value: str) -> str:
    return "Yes" if str(value).strip().lower() in {"yes", "y", "true", "1", "source"} else "No"


def _table_name(title: str) -> str:
    return "".join(ch for ch in title.title() if ch.isalnum())[:20] or "Results"
